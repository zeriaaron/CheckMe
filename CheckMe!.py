# Estimated Start : December 20, 2021
# Ended : December 31, 2021 2:00 am (Version 1.0.0.0)
# Total lines of codes: 1446 lines -unorganized
# Estimated total lines of codes: 2000+ lines -organized

from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QWidget
from PyQt5.QtCore import QTimer
from PyQt5 import uic
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
import sys


class Help(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi("Ui/help.ui", self)

        self.setFixedWidth(400)
        self.setFixedHeight(300)

        self.url_link_review = "<a href=\"https://docs.google.com/forms/d/e/1FAIpQLSe4Fjr0Mtp7dnSwSJ0Cuc96cVC0" \
                               "KhDSOqnJ8dEEgamRvS4naQ/viewform?usp=pp_url\"> <font face=Ms Shell Dlg size=5 " \
                               "color=green> Software Product Review</font> </a>"

        self.url_link_docu = "<a href=\"https://raveworks.github.io\"> " \
                             "<font face=Ms Shell Dlg size=5 color=green> " \
                             "CheckMe! Documentation</font> </a>"

        self.documentation_label.setText(self.url_link_docu)
        self.review_label.setText(self.url_link_review)


class Credits(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi("Ui/credits.ui", self)

        self.setFixedWidth(401)
        self.setFixedHeight(500)


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        uic.loadUi("Ui/check.ui", self)
        self.start_code()

    def start_code(self):
        # Non-resizable ui
        self.setFixedWidth(549)
        self.setFixedHeight(469)

        # Set timer_1 for update continuously
        # key answers
        self.timer_1 = QTimer(self)
        self.timer_1.timeout.connect(self.key_answers_timer)
        self.timer_1.start(0)

        # student answer
        self.timer_2 = QTimer(self)
        self.timer_2.timeout.connect(self.student_answer_timer)
        self.timer_2.start(0)

        # count of student id and student answer
        self.timer_3 = QTimer(self)
        self.timer_3.timeout.connect(self.count_student_id_answer_timer)
        self.timer_3.start(0)

        # Seting the starting current Tab widget
        self.tabWidget.setCurrentWidget(self.tab)

        # Visible off for Invalid Input
        # Tab 1
        self.invalid_count_students.setVisible(False)
        self.invalid_count_key_answers.setVisible(False)
        self.empty_count_students.setVisible(False)
        self.empty_count_key_answers.setVisible(False)

        # Tab 2
        self.invalid_key_answers.setVisible(False)
        self.empty_key_answers.setVisible(False)
        self.invalid_student_id.setVisible(False)
        self.invalid_student_answers.setVisible(False)
        self.empty_student_answers.setVisible(False)
        self.same_student_id.setVisible(False)

        # Tab 3
        self.invalid_confirm.setVisible(False)

        # Tab 4
        self.change_name_button_7.setEnabled(False)

        # Button connections
        self.enter_button.clicked.connect(self.enter_1)
        self.enter_button_2.clicked.connect(self.enter_2)
        self.confirm_button_3.clicked.connect(self.confirm)
        self.remove_button_4.clicked.connect(self.remove)
        self.browse_button_5.clicked.connect(self.browse)
        self.lock_name_button_6.clicked.connect(self.lock)
        self.change_name_button_7.clicked.connect(self.change)
        self.horizontal_button_7.clicked.connect(self.hori)
        self.vertical_button_8.clicked.connect(self.verti)
        self.save_file_button.clicked.connect(self.save_file)
        self.credits_button.clicked.connect(self.credits)
        self.help_button.clicked.connect(self.help)

        # Temporary disabling buttons
        self.enter_button_2.setEnabled(False)
        self.confirm_button_3.setEnabled(False)
        self.remove_button_4.setEnabled(False)
        self.horizontal_button_7.setEnabled(False)
        self.vertical_button_8.setEnabled(False)
        self.save_file_button.setEnabled(False)

        # Temporary disabling line edit tab_3
        self.key_ans_value.setEnabled(False)

        # Constant disabling line edit tab_4
        self.browse_entry.setEnabled(False)

        # Temporary disabling style sheet
        self.vertical_label.setStyleSheet("")
        self.horizontal_label.setStyleSheet("")

        # List
        self.student_id_list = []
        self.student_answers_list = []

        # Duplicates
        self.dup_student_id_list = []

        # Disabling tab temporarily
        self.tabWidget.setTabEnabled(0, True)
        self.tabWidget.setTabEnabled(1, False)
        self.tabWidget.setTabEnabled(2, False)
        self.tabWidget.setTabEnabled(3, False)

        # settings credits ui to None
        self.window = None
        self.window_2 = None

    # Timers
    def key_answers_timer(self):
        self.type_count_key_answers_label.setText(str(len(self.key_answers_entry.text())))

    def student_answer_timer(self):
        self.type_count_student_answer_label.setText(str(len(self.student_answers_entry.text())))

    def count_student_id_answer_timer(self):
        self.count_id_answer_list.setText("Count: {}".format(self.student_id_listWidget.count()))

    # Credit button
    def credits(self):
        if self.window is None:
            self.window = Credits()
        self.window.show()

    # Help button
    def help(self):
        if self.window_2 is None:
            self.window_2 = Help()
        self.window_2.show()

    # Push Buttons
    def enter_1(self):
        # Count of students
        if self.count_of_students_entry.text().strip() == '':
            self.invalid_count_students.setVisible(False)
            self.empty_count_students.setVisible(True)
        elif not self.count_of_students_entry.text().isdigit():
            self.invalid_count_students.setVisible(True)
            self.empty_count_students.setVisible(False)
        elif self.count_of_students_entry.text().isdigit():
            self.count_students_list = [int(x) for x in str(self.count_of_students_entry.text())]
            if self.count_students_list[0] == 0:
                self.invalid_count_students.setVisible(True)
                self.empty_count_students.setVisible(False)
            else:
                self.invalid_count_students.setVisible(False)
                self.empty_count_students.setVisible(False)
        else:
            self.invalid_count_students.setVisible(False)
            self.empty_count_students.setVisible(False)
            self.count_students = int(self.count_of_students_entry.text())

        # Count of Key Answers
        if self.count_key_answers_entry.text().strip() == '':
            self.invalid_count_key_answers.setVisible(False)
            self.empty_count_key_answers.setVisible(True)
        elif not self.count_key_answers_entry.text().isdigit():
            self.invalid_count_key_answers.setVisible(True)
            self.empty_count_key_answers.setVisible(False)
        elif self.count_key_answers_entry.text().isdigit():
            self.count_key_answers_list = [int(x) for x in str(self.count_key_answers_entry.text())]
            if self.count_key_answers_list[0] == 0:
                self.invalid_count_key_answers.setVisible(True)
                self.empty_count_key_answers.setVisible(False)
            else:
                self.invalid_count_key_answers.setVisible(False)
                self.empty_count_key_answers.setVisible(False)
        else:
            self.invalid_count_key_answers.setVisible(False)
            self.empty_count_key_answers.setVisible(False)
            self.count_key_answers = int(self.count_key_answers_entry.text())

        # Setting tab widget
        if self.count_of_students_entry.text().isdigit() and self.count_key_answers_entry.text().isdigit() \
                and self.count_students_list[0] != 0 and self.count_key_answers_list[0] != 0:
            self.tabWidget.setCurrentWidget(self.tab_2)
            self.enter_button_2.setEnabled(True)

            # Enabling the tab widget
            self.tabWidget.setTabEnabled(0, False)
            self.tabWidget.setTabEnabled(1, True)
            self.tabWidget.setTabEnabled(2, True)

    def enter_2(self):
        # Key answers, student id, and student's answer

        # Key answers
        if len(self.key_answers_entry.text()) != int(self.count_key_answers_entry.text()) and \
                self.key_answers_entry.text().strip():
            self.invalid_key_answers.setVisible(True)
            self.empty_key_answers.setVisible(False)

        elif not self.key_answers_entry.text().strip():
            self.empty_key_answers.setVisible(True)
            self.invalid_key_answers.setVisible(False)

        elif len(self.key_answers_entry.text()) == int(self.count_key_answers_entry.text()):
            self.invalid_key_answers.setVisible(False)
            self.empty_key_answers.setVisible(False)

        # Student ID
        if not self.student_id_entry.text().strip():
            self.invalid_student_id.setVisible(True)
            self.same_student_id.setVisible(False)
        elif self.student_id_entry.text().strip():
            self.invalid_student_id.setVisible(False)
            self.same_student_id.setVisible(False)

        # Student answers
        if len(self.student_answers_entry.text()) != int(self.count_key_answers_entry.text()) and \
                self.student_answers_entry.text().strip():
            self.invalid_student_answers.setVisible(True)
            self.empty_student_answers.setVisible(False)

        elif not self.student_answers_entry.text().strip():  # empty
            self.empty_student_answers.setVisible(True)
            self.invalid_student_answers.setVisible(False)

        elif len(self.student_answers_entry.text()) == int(self.count_key_answers_entry.text()):
            self.invalid_student_answers.setVisible(False)
            self.empty_student_answers.setVisible(False)

        # If all true
        if len(self.key_answers_entry.text()) == int(self.count_key_answers_entry.text()) and \
                self.student_id_entry.text().strip() and len(self.student_answers_entry.text()) == \
                int(self.count_key_answers_entry.text()):
            self.invalid_key_answers.setVisible(False)
            self.empty_key_answers.setVisible(False)
            self.invalid_student_id.setVisible(False)
            self.invalid_student_answers.setVisible(False)
            self.empty_student_answers.setVisible(False)

            # Student id duplicates
            for self.i in range(self.student_id_listWidget.count()):
                if self.student_id_listWidget.item(self.i).text() != '':
                    self.dup_student_id_list.append(self.student_id_listWidget.item(self.i).text())

            if not self.student_id_entry.text().upper() in self.dup_student_id_list:
                self.student_id_listWidget.addItem(self.student_id_entry.text().upper())

                # Getting and prompting key answers
                self.key_answers = self.key_answers_entry.text().upper()
                self.key_ans_value.setText(self.key_answers)

                # Appending student's id and answer into listWidget
                self.student_answers_listWidget.addItem(self.student_answers_entry.text().upper())

                # Clearing the entries
                self.student_id_entry.clear()
                self.student_answers_entry.clear()

                self.confirm_button_3.setEnabled(True)
                self.remove_button_4.setEnabled(True)

                self.confirm_button_3.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                                    "border-bor_rad: 10px; } "
                                                    "QPushButton:hover:!pressed "
                                                    "{ background-color: rgb(255, 255, 0); }")

                if self.student_id_listWidget.count() == int(self.count_of_students_entry.text()) and \
                        self.student_answers_listWidget.count() == int(self.count_of_students_entry.text()):
                    self.student_id_entry.setEnabled(False)
                    self.student_answers_entry.setEnabled(False)
                    self.enter_button_2.setEnabled(False)
                    self.key_answers_entry.setEnabled(False)

            else:
                self.same_student_id.setVisible(True)

        # set current widget
        if self.student_id_listWidget.count() == int(self.count_of_students_entry.text()) and \
            self.student_answers_listWidget.count() == int(self.count_of_students_entry.text()):
            # Set current tab widget
            self.tabWidget.setCurrentWidget(self.tab_3)

    def remove(self):
        self.uni_index = self.student_id_listWidget.currentIndex().row()
        self.student_id_listWidget.takeItem(self.uni_index)
        self.student_answers_listWidget.takeItem(self.uni_index)

        if self.student_id_listWidget.count() != int(self.count_of_students_entry.text()) and \
                self.student_answers_listWidget.count() != int(self.count_of_students_entry.text()):
            self.student_id_entry.setEnabled(True)
            self.student_answers_entry.setEnabled(True)
            self.enter_button_2.setEnabled(True)
            self.key_answers_entry.setEnabled(True)

            # Student id duplicates
            self.dup_student_id_list.clear()
            for self.i in range(self.student_id_listWidget.count()):
                if self.student_id_listWidget.item(self.i).text() != '':
                    self.dup_student_id_list.append(self.student_id_listWidget.item(self.i).text())

        if self.student_id_listWidget.count() == 0:
            pass

    def confirm(self):
        # False condition
        if self.student_id_listWidget.count() != int(self.count_of_students_entry.text()) and \
                self.student_answers_listWidget.count() != int(self.count_of_students_entry.text()):
            # invalid_confirm timer

            # CONFUSION: singleShot is continuously checking the setVisible(True) and not the function, the function
            # is like the happenings in default mode.
            self.timer_4 = QTimer(self)
            self.timer_4.singleShot(3000, self.invalid_confirm_button)
            self.invalid_confirm.setVisible(True)

        # Correct condition
        elif self.student_id_listWidget.count() == int(self.count_of_students_entry.text()) and \
                self.student_answers_listWidget.count() == int(self.count_of_students_entry.text()):
            # Appending student id to a list
            for self.i in range(self.student_id_listWidget.count()):
                if self.student_id_listWidget.item(self.i).text() != '':
                    self.student_id_list.append(self.student_id_listWidget.item(self.i).text())

            # Appending student answer to a list
            for self.i in range(self.student_answers_listWidget.count()):
                if self.student_answers_listWidget.item(self.i).text() != '':
                    self.student_answers_list.append(self.student_answers_listWidget.item(self.i).text())

            self.dict_id_ans = {}
            # self.student_id_list = []
            # self.student_answers_list = []
            for self.student_id in self.student_id_list:
                for self.student_answer in self.student_answers_list:
                    self.dict_id_ans[self.student_id] = self.student_answer
                    self.student_answers_list.remove(self.student_answer)
                    break

            self.dict_id_ans_items = self.dict_id_ans.items()
            self.sorted_dict_id_ans = sorted(self.dict_id_ans_items)
            self.dict_again = dict(self.sorted_dict_id_ans)

            # List of sorted student id and student answers
            self.sort_student_id = list(self.dict_again.keys())
            self.sort_student_ans = list(self.dict_again.values())

            # Scores of students
            self.score_list = []

            # Check Algorithm
            for i in range(int(self.count_of_students_entry.text())):
                self.sum = 0
                for j in range(int(self.count_key_answers_entry.text())):
                    self.answer = 0
                    if self.key_answers[j] == self.sort_student_ans[i][j]:
                        self.answer += 1
                    self.sum = self.sum + self.answer
                self.score_list.append(self.sum)

            # set enabled tab widget
            self.tabWidget.setTabEnabled(0, False)
            self.tabWidget.setTabEnabled(1, False)
            self.tabWidget.setTabEnabled(2, False)
            self.tabWidget.setTabEnabled(3, True)

            # Setting the current tab
            self.tabWidget.setCurrentWidget(self.tab_4)

            # Disabling the confirm button and remove selected item button
            self.confirm_button_3.setEnabled(False)
            self.remove_button_4.setEnabled(False)

            # Setting the style sheet for confirm button
            self.confirm_button_3.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0);"
                                                "border-bor_rad: 10px; }")

    def invalid_confirm_button(self):
        # default mode
        self.invalid_confirm.setVisible(False)

    def browse(self):
        self.file_path = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
        self.browse_entry.setText(self.file_path)

    def change(self):
        self.file_name_entry.setEnabled(True)
        self.lock_name_button_6.setEnabled(True)
        self.change_name_button_7.setEnabled(False)

        self.change_name_button_7.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0);"
                                                "border-bor_rad: 10px; }")

        self.lock_name_button_6.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                              "border-bor_rad: 10px; }"
                                              "QPushButton:hover:!pressed { background-color: rgb(255, 255, 0); }")

        # Disabling buttons temporarily when changed
        self.horizontal_button_7.setEnabled(False)
        self.vertical_button_8.setEnabled(False)

        # Clearing the line edit if it is filled
        self.file_name_entry.clear()

        # Disabling the style sheet to labels temporarily
        self.vertical_label.setStyleSheet("")
        self.horizontal_label.setStyleSheet("")

    def lock(self):
        self.file_name_entry.setEnabled(False)
        self.lock_name_button_6.setEnabled(False)
        self.change_name_button_7.setEnabled(True)

        self.lock_name_button_6.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0); "
                                              "border-bor_rad: 10px; }")

        self.change_name_button_7.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0);"
                                                "border-bor_rad: 10px; }"
                                                "QPushButton:hover:!pressed { background-color: rgb(255, 255, 0); }")

        if self.browse_entry.text() != '' and self.file_name_entry.text() != '':
            self.horizontal_button_7.setEnabled(True)
            self.vertical_button_8.setEnabled(True)

            # Setting the style sheet for vertical and horizontal layouts
            self.vertical_label.setStyleSheet("QLabel:hover {background-color: rgb(255, 255, 0); border-bor_rad: 10px;}")
            self.horizontal_label.setStyleSheet("QLabel:hover {background-color: rgb(255, 255, 0); "
                                                "border-bor_rad: 10px;}")

    def hori(self):
        # Excel Creation
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Check Data Results"

        self.horizontal_button_7.setEnabled(False)
        self.horizontal_button_7.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0); "
                                               "border-bor_rad: 10px; }")

        self.vertical_button_8.setEnabled(True)
        self.vertical_button_8.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); border-bor_rad: 10px; } "
                                             "QPushButton:hover:!pressed { background-color: rgb(255, 255, 0); }")
        self.vertical_label.setStyleSheet("")
        self.horizontal_label.setStyleSheet("QLabel:hover { background-color: rgb(255, 255, 0);"
                                            "border-bor_rad: 10px; }")

        self.save_file_button.setEnabled(True)

        self.ws.append(["Student IDs"] + self.sort_student_id)
        self.ws.append(["Scores"] + self.score_list)

        for self.col in range(1, len(self.sort_student_id) + 2):
            self.col_letter = get_column_letter(self.col)
            self.ws["{}1".format(self.col_letter)].alignment = Alignment(horizontal="center")
            self.ws["{}2".format(self.col_letter)].alignment = Alignment(horizontal="center")

        self.ws.column_dimensions['A'].width = 15

    def verti(self):
        # Excel Creation
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Check Data Results"

        self.vertical_button_8.setEnabled(False)
        self.vertical_button_8.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0); "
                                             "border-bor_rad: 10px; }")
        self.horizontal_button_7.setEnabled(True)
        self.horizontal_button_7.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); border-bor_rad: 10px; } "
                                               "QPushButton:hover:!pressed { background-color: rgb(255, 255, 0); }")

        self.horizontal_label.setStyleSheet("")
        self.vertical_label.setStyleSheet("QLabel:hover { background-color: rgb(255, 255, 0);"
                                          "border-bor_rad: 10px; }")

        self.save_file_button.setEnabled(True)

        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 15

        self.ws["A1"] = "Student IDs"
        self.ws["B1"] = "Scores"

        for i in range(len(self.sort_student_id)):
            self.ws["A{}".format(i + 2)] = self.sort_student_id[i]
            self.ws["A{}".format(i + 2)].alignment = Alignment(horizontal="center")

        for i in range(len(self.score_list)):
            self.ws["B{}".format(i + 2)] = self.score_list[i]
            self.ws["B{}".format(i + 2)].alignment = Alignment(horizontal="center")

        self.ws["A1"].alignment = Alignment(horizontal="center")
        self.ws["B1"].alignment = Alignment(horizontal="center")

    def save_file(self):
        if self.browse_entry.text()[-1] == '/':
            self.wb.save("{}{}.xlsx".format(self.browse_entry.text(), self.file_name_entry.text()))
            self.wb.close()
            self.save_file_button.setEnabled(False)
            self.tabWidget.setTabEnabled(3, False)

            # set current tab
            self.tabWidget.setCurrentWidget(self.tab)

            # Clearing all Widgets
            self.student_id_listWidget.clear()
            self.student_answers_listWidget.clear()
            self.key_ans_value.clear()
            self.key_answers_entry.clear()
            self.student_id_entry.clear()
            self.student_answers_entry.clear()
            self.count_of_students_entry.clear()
            self.count_key_answers_entry.clear()
            self.browse_entry.clear()
            self.file_name_entry.clear()
            self.dup_student_id_list.clear()
            self.student_id_list.clear()
            self.student_answers_list.clear()
            self.sort_student_id.clear()
            self.sort_student_ans.clear()
            self.score_list.clear()
            self.dict_id_ans.clear()
            self.sorted_dict_id_ans.clear()
            self.dict_again.clear()

            # Enabling the buttons
            self.key_answers_entry.setEnabled(True)
            self.student_id_entry.setEnabled(True)
            self.student_answers_entry.setEnabled(True)

            self.file_name_entry.setEnabled(True)
            self.lock_name_button_6.setEnabled(True)
            self.change_name_button_7.setEnabled(False)

            # Style sheet for change and lock
            self.change_name_button_7.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0); "
                                                    "border-bor_rad: 10px; } QPushButton:hover:!pressed { "
                                                    "background-color: rgb(255, 255, 0); }")
            self.lock_name_button_6.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                                  "border-bor_rad: 10px; } QPushButton:hover:!pressed { "
                                                  "background-color: rgb(255, 255, 0); } ")

            # pop list
            self.dup_student_id_list.clear()

            # hori
            self.horizontal_label.setStyleSheet("QLabel:hover { background-color: rgb(255, 255, 0); "
                                                "border-bor_rad: 10px; }")
            self.horizontal_button_7.setEnabled(False)
            self.horizontal_button_7.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                                   "border-bor_rad: 10px; } QPushButton:hover:!pressed "
                                                   "{ background-color: rgb(255, 255, 0); }")

            # verti
            self.vertical_label.setStyleSheet("QLabel:hover {background-color: rgb(255, 255, 0); "
                                              "border-bor_rad: 10px;}")
            self.vertical_button_8.setEnabled(False)
            self.vertical_button_8.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); border-bor_rad: 10px; "
                                                 "} QPushButton:hover:!pressed { "
                                                 "background-color: rgb(255, 255, 0); }")

            # Disabling tab temporarily
            self.tabWidget.setTabEnabled(0, True)
            self.tabWidget.setTabEnabled(1, False)
            self.tabWidget.setTabEnabled(2, False)
            self.tabWidget.setTabEnabled(3, False)

        else:
            self.wb.save("{}/{}.xlsx".format(self.browse_entry.text(), self.file_name_entry.text()))
            self.wb.close()
            self.save_file_button.setEnabled(False)
            self.tabWidget.setTabEnabled(3, False)

            # set current tab
            self.tabWidget.setCurrentWidget(self.tab)

            # Clearing all Widgets
            self.student_id_listWidget.clear()
            self.student_answers_listWidget.clear()
            self.key_ans_value.clear()
            self.key_answers_entry.clear()
            self.student_id_entry.clear()
            self.student_answers_entry.clear()
            self.count_of_students_entry.clear()
            self.count_key_answers_entry.clear()
            self.browse_entry.clear()
            self.file_name_entry.clear()
            self.dup_student_id_list.clear()
            self.student_id_list.clear()
            self.student_answers_list.clear()
            self.sort_student_id.clear()
            self.sort_student_ans.clear()
            self.score_list.clear()
            self.dict_id_ans.clear()
            self.sorted_dict_id_ans.clear()
            self.dict_again.clear()

            # Enabling and Disabling the buttons
            self.key_answers_entry.setEnabled(True)
            self.student_id_entry.setEnabled(True)
            self.student_answers_entry.setEnabled(True)

            self.file_name_entry.setEnabled(True)
            self.lock_name_button_6.setEnabled(True)
            self.change_name_button_7.setEnabled(False)

            # Style sheet for change and lock
            self.change_name_button_7.setStyleSheet("QPushButton { background-color: rgb(255, 255, 0); "
                                                    "border-bor_rad: 10px; } QPushButton:hover:!pressed { "
                                                    "background-color: rgb(255, 255, 0); }")
            self.lock_name_button_6.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                                  "border-bor_rad: 10px; } QPushButton:hover:!pressed { "
                                                  "background-color: rgb(255, 255, 0); } ")

            # pop list
            self.dup_student_id_list.clear()

            # hori
            self.horizontal_label.setStyleSheet("QLabel:hover { background-color: rgb(255, 255, 0); "
                                                "border-bor_rad: 10px; }")
            self.horizontal_button_7.setEnabled(False)
            self.horizontal_button_7.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); "
                                                   "border-bor_rad: 10px; } QPushButton:hover:!pressed "
                                                   "{ background-color: rgb(255, 255, 0); }")

            # verti
            self.vertical_label.setStyleSheet("QLabel:hover {background-color: rgb(255, 255, 0); "
                                              "border-bor_rad: 10px;}")
            self.vertical_button_8.setEnabled(False)
            self.vertical_button_8.setStyleSheet("QPushButton { background-color: rgb(0, 255, 0); border-bor_rad: 10px; "
                                                 "} QPushButton:hover:!pressed { "
                                                 "background-color: rgb(255, 255, 0); }")

            # Disabling tab temporarily
            self.tabWidget.setTabEnabled(0, True)
            self.tabWidget.setTabEnabled(1, False)
            self.tabWidget.setTabEnabled(2, False)
            self.tabWidget.setTabEnabled(3, False)


app = QApplication(sys.argv)
root = MainWindow()
root.show()
sys.exit(app.exec_())
